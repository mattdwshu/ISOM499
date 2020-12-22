import re
from pprint import pprint
import json
import openpyxl
import requests
from yelpapi import YelpAPI

api_key = "INSERT API KEY HERE"
headers = {f'Authorization': f'Bearer {api_key}'}

# Searching in the yelpAPI for my restaurant chain, Sweet Hut
yelp_api = YelpAPI(api_key)
resp_sweet = yelp_api.search_query(term='Sweet Hut', location='atlanta, ga', sort_by='rating', limit=1)

# Getting the business ID for Sweet Hut from the YelpAPI
for i in resp_sweet['businesses']:
    biz_id = i['id']

# Getting JSON data for reviews
sweetURL = f'https://api.yelp.com/v3/businesses/{biz_id}/reviews'
response1 = requests.get(sweetURL, headers=headers)
sweet_reviews = json.loads(response1.text)
pprint(sweet_reviews)
r = sweet_reviews['reviews']
print('Sweet Hut Reviews:')
print(r[0]['text'], '\n')
print(r[1]['text'], '\n')
print(r[2]['text'], '\n')

# Now going through word frequencies. I concatenate the reviews to make things easier for me
sweetrev = (r[0]['text'] + ' ' + r[1]['text'] + ' ' + r[2]['text'])

# I first remove punctuation because that can mess up the word count
sweet1rev = re.sub(r'[^\w\s]', '', sweetrev)

# Then I turn the string into a dictionary and make everything lowercase
sweetdata = sweet1rev.lower().split()
sweetWordFreq = {}

# Words chosen: location, parking, love, tea, donut, chicken,
# Because these are all 5 star reviews, I'm assuming a higher word frequency is better.
# I also save the frequencies as a variable so I can use it later when exporting
print('Word Frequencies:')
for word in sweetdata:
    if word in sweetWordFreq:
        sweetWordFreq[word] += 1
    else:
        sweetWordFreq[word] = 1
loc = 0
for loca in sweetWordFreq:
    if loca == 'location':
        print(loca, ':', sweetWordFreq[loca])
        loc = sweetWordFreq[loca]

par = 0
for park in sweetWordFreq:
    if park == 'parking':
        print(park, ':', sweetWordFreq[park])
        par = sweetWordFreq[park]

lov = 0
for love in sweetWordFreq:
    if love == 'love':
        print(love, ':', sweetWordFreq[love])
        lov = sweetWordFreq[love]

tea = 0
for t in sweetWordFreq:
    if t == 'tea':
        print(t, ':', sweetWordFreq[t])
        tea = sweetWordFreq[t]

d = 0
for don in sweetWordFreq:
    if don == 'donut':
        print(don, ':', sweetWordFreq[don])
        d = sweetWordFreq[don]

ch = 0
for chik in sweetWordFreq:
    if chik == 'chicken':
        print(chik, sweetWordFreq[chik])
        ch = sweetWordFreq[chik]

# exporting to csv file
wb = openpyxl.Workbook()
ws = wb.create_sheet(title = 'Word Frequencies')
ws['A1'] = 'Sweet Hut Words:'
ws['A2'] = 'Location'
ws['A3'] = 'Parking'
ws['A4'] = 'Love'
ws['A5'] = 'Tea'
ws['A6'] = 'Donut'
ws['A7'] = 'Chicken'
ws['B2'] = loc
ws['B3'] = par
ws['B4'] = lov
ws['B5'] = tea
ws['B6'] = d
ws['B7'] = ch

print('\n','************','\n')

# Now I'm going through the same process with my competitors
resp_comp = yelp_api.search_query(term='Bubble tea', location='atlanta, ga', sort_by='rating', limit=1)
for i in resp_comp['businesses']:
    comp_id = i['id']
compURL = f'https://api.yelp.com/v3/businesses/{comp_id}/reviews'
response2 = requests.get(compURL, headers=headers)
comp_reviews = json.loads(response2.text)
c = comp_reviews['reviews']
print('Competitor Reviews:')
print(c[0]['text'], '\n')
print(c[1]['text'], '\n')
print(c[2]['text'], '\n')

# Same deal with the competitor reviews

comprev = (c[0]['text'] + ' ' + c[1]['text'] + ' ' + c[2]['text'])
comp1rev = re.sub(r'[^\w\s]', '', comprev)
compdata = comp1rev.lower().split()
compWordFreq = {}

# I'm choosing different words for the competitor based on what I saw from the yelp reviews
# tea, online, fast, employees
print('Word Frequencies:')
for word in compdata:
    if word in compWordFreq:
        compWordFreq[word] += 1
    else:
        compWordFreq[word] = 1

tea1 = 0
for t in compWordFreq:
    if t == 'tea':
        print(t, ':', compWordFreq[t])
        tea1 = compWordFreq[t]

onl = 0
for onli in compWordFreq:
    if onli == 'online':
        print(onli, ':', compWordFreq[onli])
        onl = compWordFreq[onli]

f = 0
for fast in compWordFreq:
    if fast == 'fast':
        print(fast, ':', compWordFreq[fast])
        f = compWordFreq[fast]

emp = 0
for empl in compWordFreq:
    if empl == 'employees':
        print(empl, ':', compWordFreq[empl])
        emp = compWordFreq[empl]

ws['A9'] = 'Competitor Words:'
ws['A10'] = 'Tea'
ws['A11'] = 'Online'
ws['A12'] = 'Fast'
ws['A13'] = 'empl'
ws['B10'] = tea1
ws['B11'] = onl
ws['B12'] = f
ws['B13'] = emp

wb.save(filename = "word_frequencies.csv")